from __future__ import annotations

import csv
import html
import re
import traceback
from collections import Counter, defaultdict
from copy import deepcopy
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import Workbook

from config import (
    APP_SECRET_EXPIRY_DATE,
    APP_SECRET_WARNING_DAYS,
    BROKEN_SCRAPERS_FILE,
    CATEGORY_RULES_FILE,
    CSV_FILE,
    DISABLE_BLACKLIST,
    EMAIL_CONFIG,
    EMAIL_ENABLED,
    EXCEL_FILE,
    REMOVED_JOBS_FILE,
    RUN_SCRAPERS,
    TITLE_BLACKLIST_FILE,
    TITLE_OVERRIDES_FILE,
    UNRELATED_JOBS_FILE,
)
from filters.blacklist import is_title_blacklisted, load_title_blacklist
from filters.job_classifier import JobClassifier
from graph_mailer import send_report_email_graph

# Scraper imports
from scrapers.ardis import ArdisScraper
from scrapers.astrid import AstridScraper
from scrapers.battmobility import BattMobilityScraper
from scrapers.bizztalent import BizzTalentScraper
from scrapers.cheops import CheopsScraper
from scrapers.cobral import CobralScraper
from scrapers.colruytgroup import ColruytGroupScraper
from scrapers.crelan import CrelanScraper
from scrapers.dataclean import DataCleanScraper
from scrapers.decathlon import DecathlonScraper
from scrapers.district09 import District09Scraper
from scrapers.durabrik import DurabrikScraper
from scrapers.equalminds import EqualMindsScraper
from scrapers.evara import EvaraScraper
from scrapers.fluxys import FluxysScraper
from scrapers.harveynash import HarveyNashScraper
from scrapers.hays import HaysScraper
from scrapers.hieronymus import HieronymusScraper
from scrapers.hogent import HOGentScraper
from scrapers.hr_planet import HRPlanetScraper
from scrapers.infrabel import InfrabelScraper
from scrapers.it_planet import ITPlanetScraper
from scrapers.itprovider import ITProviderScraper
from scrapers.itsgroup import ITSGroupScraper
from scrapers.logitechnic import LogiTechnicScraper
from scrapers.meatandmore import MeatAndMoreScraper
from scrapers.oost_vlaanderen import OostVlaanderenScraper
from scrapers.projinit import ProjinitScraper
from scrapers.resolvus import ResolvusScraper
from scrapers.signpost import SignpostScraper
from scrapers.simac import SimacScraper
from scrapers.solidaris import SolidarisScraper
from scrapers.splendit import SplendITScraper
from scrapers.synamedia import SynamediaScraper
from scrapers.taborgroep import TaborGroepScraper
from scrapers.talencia import TalenciaScraper
from scrapers.televic import TelevicScraper
from scrapers.uilenspel import UilenspelScraper
from scrapers.upgrade_estate import UpgradeEstateScraper
from scrapers.vdkbank import VDKScraper
from scrapers.verdon import VerdonScraper
from scrapers.vlaanderen_connect import VlaanderenConnectScraper
from scrapers.volvocars import VolvoCarsScraper
from scrapers.xcare import XCareScraper
from scrapers.xelor import XelorScraper

FIELDNAMES = [
    "source",
    "title",
    "url",
    "location",
    "location_is_guess",
    "first_seen_date",
    "last_seen_date",
    "is_active",
    "category",
    "subcategory",
]


def enrich_job_with_classification(
    job: Dict[str, str],
    classifier: JobClassifier,
) -> Dict[str, str]:
    result = classifier.classify_job_title(job.get("title", ""))

    enriched = dict(job)
    enriched["category"] = result.category
    enriched["subcategory"] = result.subcategory
    return enriched


def load_existing_jobs(
    csv_file: Path,
    classifier: JobClassifier,
) -> Dict[Tuple[str, str], Dict[str, str]]:
    jobs: Dict[Tuple[str, str], Dict[str, str]] = {}

    if not csv_file.exists():
        return jobs

    with csv_file.open("r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            key = (row["source"], row["url"])

            if "location" not in row:
                row["location"] = ""
            if "location_is_guess" not in row:
                row["location_is_guess"] = "False"
            if "category" not in row:
                row["category"] = ""
            if "subcategory" not in row:
                row["subcategory"] = ""

            if not row["category"]:
                row = enrich_job_with_classification(row, classifier)

            jobs[key] = row

    return jobs


def get_active_jobs(
    jobs: Dict[Tuple[str, str], Dict[str, str]],
) -> Dict[Tuple[str, str], Dict[str, str]]:
    return {
        key: row
        for key, row in jobs.items()
        if row.get("is_active") == "True"
    }


def save_jobs_csv(
    csv_file: Path,
    jobs: Dict[Tuple[str, str], Dict[str, str]],
) -> None:
    rows = sorted(
        jobs.values(),
        key=lambda row: (row["source"], row["title"], row["url"]),
    )

    with csv_file.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(rows)


def auto_fit_worksheet_columns(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)

        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)


def save_jobs_excel(
    output_file: Path,
    jobs: Dict[Tuple[str, str], Dict[str, str]],
) -> None:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    jobs_by_source: dict[str, list[Dict[str, str]]] = defaultdict(list)

    for row in jobs.values():
        source = row.get("source", "unknown")
        jobs_by_source[source].append(row)

    for source, rows in sorted(jobs_by_source.items(), key=lambda x: x[0].lower()):
        sheet_name = source[:31] if source else "unknown"
        ws = wb.create_sheet(title=sheet_name)

        ws.append(FIELDNAMES)

        rows = sorted(rows, key=lambda row: (row.get("title", ""), row.get("url", "")))

        for row in rows:
            ws.append([row.get(field, "") for field in FIELDNAMES])

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        auto_fit_worksheet_columns(ws)

    if not wb.sheetnames:
        ws = wb.create_sheet(title="jobs")
        ws.append(FIELDNAMES)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        auto_fit_worksheet_columns(ws)

    wb.save(output_file)


def save_unrelated_jobs_txt(
    output_file: Path,
    jobs: Dict[Tuple[str, str], Dict[str, str]],
) -> None:
    unrelated_rows = sorted(
        [
            row
            for row in jobs.values()
            if row.get("is_active") == "True" and row.get("category") == "unrelated"
        ],
        key=lambda row: (row.get("source", ""), row.get("title", "")),
    )

    with output_file.open("w", encoding="utf-8") as f:
        if not unrelated_rows:
            f.write("No active unrelated jobs found.\n")
            return

        for row in unrelated_rows:
            f.write(f"source: {row.get('source', '')}\n")
            f.write(f"title: {row.get('title', '')}\n")
            f.write(f"url: {row.get('url', '')}\n")
            f.write(f"location: {row.get('location', '')}\n")
            f.write(f"subcategory: {row.get('subcategory', '')}\n")
            f.write("\n")


def save_broken_scrapers_txt(
    output_file: Path,
    broken_scrapers: List[tuple[str, str]],
) -> None:
    with output_file.open("w", encoding="utf-8") as f:
        if not broken_scrapers:
            f.write("No broken scrapers.\n")
            return

        for scraper_name, error_message in broken_scrapers:
            f.write(f"{scraper_name}: {error_message}\n")


def save_removed_jobs_txt(
    output_file: Path,
    removed_jobs: List[Dict[str, str]],
) -> None:
    with output_file.open("w", encoding="utf-8") as f:
        if not removed_jobs:
            f.write("No jobs were removed this run.\n")
            return

        for row in sorted(
            removed_jobs,
            key=lambda r: (r.get("source", ""), r.get("title", "")),
        ):
            f.write(f"source: {row.get('source', '')}\n")
            f.write(f"title: {row.get('title', '')}\n")
            f.write(f"url: {row.get('url', '')}\n")
            f.write(f"location: {row.get('location', '')}\n")
            f.write(f"last_seen_date: {row.get('last_seen_date', '')}\n")
            f.write("\n")


def remove_blacklisted_existing_jobs(
    existing_jobs: Dict[Tuple[str, str], Dict[str, str]],
    title_blacklist,
) -> Dict[Tuple[str, str], Dict[str, str]]:
    cleaned_jobs: Dict[Tuple[str, str], Dict[str, str]] = {}

    for key, row in existing_jobs.items():
        title = row.get("title", "")
        source = row.get("source", "")

        if is_title_blacklisted(title, title_blacklist, source=source):
            continue

        cleaned_jobs[key] = row

    return cleaned_jobs


def filter_scraped_jobs(
    scraped_jobs: List[Dict[str, str]],
    title_blacklist,
    classifier: JobClassifier,
) -> List[Dict[str, str]]:
    filtered_jobs: List[Dict[str, str]] = []

    for job in scraped_jobs:
        title = job.get("title", "")
        source = job.get("source", "")

        if not DISABLE_BLACKLIST:
            if is_title_blacklisted(title, title_blacklist, source=source):
                continue

        job = enrich_job_with_classification(job, classifier)
        filtered_jobs.append(job)

    return filtered_jobs


def merge_scrape_results(
    existing_jobs: Dict[Tuple[str, str], Dict[str, str]],
    scraped_jobs: List[Dict[str, str]],
    scraped_date: str,
) -> tuple[
    Dict[Tuple[str, str], Dict[str, str]],
    List[Dict[str, str]],
    List[Dict[str, str]],
]:
    current_keys = set()
    scraped_sources = {job["source"] for job in scraped_jobs}
    new_jobs: List[Dict[str, str]] = []
    removed_jobs: List[Dict[str, str]] = []

    for job in scraped_jobs:
        key = (job["source"], job["url"])
        current_keys.add(key)

        location = job.get("location", "").strip()
        location_is_guess = str(job.get("location_is_guess", "False")).strip() or "False"

        category = job.get("category", "")
        subcategory = job.get("subcategory", "")

        if key in existing_jobs:
            existing_jobs[key]["title"] = job["title"]
            existing_jobs[key]["location"] = location
            existing_jobs[key]["location_is_guess"] = location_is_guess
            existing_jobs[key]["last_seen_date"] = scraped_date
            existing_jobs[key]["is_active"] = "True"
            existing_jobs[key]["category"] = category
            existing_jobs[key]["subcategory"] = subcategory
        else:
            new_row = {
                "source": job["source"],
                "title": job["title"],
                "url": job["url"],
                "location": location,
                "location_is_guess": location_is_guess,
                "first_seen_date": job.get("first_seen_date") or scraped_date,
                "last_seen_date": scraped_date,
                "is_active": "True",
                "category": category,
                "subcategory": subcategory,
            }
            existing_jobs[key] = new_row
            new_jobs.append(deepcopy(new_row))

    for key, row in existing_jobs.items():
        if (
            key not in current_keys
            and row["source"] in scraped_sources
            and row.get("is_active") == "True"
        ):
            row["is_active"] = "False"
            row["last_seen_date"] = scraped_date
            removed_jobs.append(deepcopy(row))

    return existing_jobs, new_jobs, removed_jobs


def count_unrelated_jobs(
    jobs: Dict[Tuple[str, str], Dict[str, str]],
) -> int:
    return sum(
        1
        for row in jobs.values()
        if row.get("is_active") == "True" and row.get("category") == "unrelated"
    )


def build_category_counts(
    jobs: Dict[Tuple[str, str], Dict[str, str]],
) -> Counter:
    active_jobs = [row for row in jobs.values() if row.get("is_active") == "True"]
    return Counter(row.get("category", "missing") for row in active_jobs)


def should_show_secret_expiry_warning() -> tuple[bool, int]:
    if not APP_SECRET_EXPIRY_DATE:
        return False, 0

    expiry_date = datetime.strptime(APP_SECRET_EXPIRY_DATE, "%Y-%m-%d").date()
    today = date.today()
    days_left = (expiry_date - today).days

    return 0 <= days_left <= APP_SECRET_WARNING_DAYS, days_left


def build_secret_expiry_warning_html() -> str:
    show_warning, days_left = should_show_secret_expiry_warning()

    if not show_warning:
        return ""

    expiry_text = html.escape(APP_SECRET_EXPIRY_DATE)

    return f"""
<table width="100%" cellpadding="12" cellspacing="0" border="0" style="margin-bottom:20px; border:3px solid #ff0000; background-color:#ffe5e5;">
<tr>
<td style="font-family: Arial, sans-serif; font-size:14px; color:#b00000;">
<b>⚠ APPLICATION SECRET EXPIRING SOON</b><br><br>
Your Microsoft Entra application secret will expire soon.<br>
Please generate a new secret in Entra and update it in <b>config.py</b>.<br><br>
Expiry date: <b>{expiry_text}</b><br>
Days remaining: <b>{days_left}</b>
</td>
</tr>
</table>
"""


def build_email_body(
    category_counts: Counter,
    new_jobs: List[Dict[str, str]],
    removed_jobs: List[Dict[str, str]],
    unrelated_count: int,
    broken_scrapers: List[tuple[str, str]],
    scraped_date: str,
) -> str:
    warning_html = build_secret_expiry_warning_html()

    def section_title(title: str) -> str:
        return f"""
<tr>
  <td style="padding:14px 0 8px 0; font-family:Arial,sans-serif; font-size:18px; font-weight:bold; color:#222222;">
    {html.escape(title)}
  </td>
</tr>
"""

    def paragraph(text: str) -> str:
        return f"""
<tr>
  <td style="padding:0 0 8px 0; font-family:Arial,sans-serif; font-size:14px; color:#222222;">
    {text}
  </td>
</tr>
"""

    def bullet_list(items: List[str]) -> str:
        if not items:
            return paragraph("None")

        rows = []
        for item in items:
            rows.append(
                f"""
<tr>
  <td style="padding:0 0 6px 18px; font-family:Arial,sans-serif; font-size:14px; color:#222222;">
    • {item}
  </td>
</tr>
"""
            )
        return "".join(rows)

    parts: List[str] = []

    parts.append("""
<html>
<body style="margin:0; padding:0; background-color:#f4f4f4;">
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="border-collapse:collapse; background-color:#f4f4f4;">
  <tr>
    <td align="center" style="padding:24px;">
      <table width="900" cellpadding="0" cellspacing="0" border="0" style="border-collapse:collapse; width:900px; max-width:900px; background-color:#ffffff; border:1px solid #dddddd;">
        <tr>
          <td style="padding:24px; font-family:Arial,sans-serif; font-size:14px; color:#222222;">
""")

    if warning_html:
        parts.append(warning_html)

    parts.append(f"""
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="border-collapse:collapse;">
  <tr>
    <td style="padding:0 0 18px 0; font-family:Arial,sans-serif; font-size:24px; font-weight:bold; color:#111111;">
      Job scraper report - {html.escape(scraped_date)}
    </td>
  </tr>
</table>
""")

    parts.append("""
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="border-collapse:collapse;">
""")

    parts.append(section_title("Category overview"))

    if category_counts:
        category_items = [
            f"<b>{html.escape(category)}</b>: {count}"
            for category, count in sorted(category_counts.items(), key=lambda x: (-x[1], x[0]))
        ]
        parts.append(bullet_list(category_items))
    else:
        parts.append(paragraph("No active jobs found."))

    parts.append(section_title("Unrelated jobs to review"))
    if unrelated_count > 0:
        parts.append(paragraph(f"<b>{unrelated_count}</b> unrelated jobs need review."))
    else:
        parts.append(paragraph("None"))

    parts.append(section_title(f"New jobs this run: {len(new_jobs)}"))
    new_job_items = [
        f"[{html.escape(row.get('source', ''))}] {html.escape(row.get('title', ''))}"
        for row in sorted(new_jobs, key=lambda r: (r.get('source', ''), r.get('title', '')))
    ]
    parts.append(bullet_list(new_job_items))

    parts.append(section_title(f"Removed jobs this run: {len(removed_jobs)}"))
    removed_job_items = [
        f"[{html.escape(row.get('source', ''))}] {html.escape(row.get('title', ''))}"
        for row in sorted(removed_jobs, key=lambda r: (r.get('source', ''), r.get('title', '')))
    ]
    parts.append(bullet_list(removed_job_items))

    parts.append(section_title(f"Broken scrapers: {len(broken_scrapers)}"))
    broken_scraper_items = [
        f"<b>{html.escape(scraper_name)}</b>: {html.escape(error_message)}"
        for scraper_name, error_message in broken_scrapers
    ]
    parts.append(bullet_list(broken_scraper_items))

    parts.append(section_title("Attached files"))
    parts.append(
        bullet_list(
            [
                "jobs.csv",
                "jobs.xlsx",
                "unrelated_jobs.txt",
                "broken_scrapers.txt",
                "removed_jobs.txt",
            ]
        )
    )

    parts.append("""
</table>
""")

    parts.append("""
          </td>
        </tr>
      </table>
    </td>
  </tr>
</table>
</body>
</html>
""")

    return "".join(parts)


def send_report_email(
    subject: str,
    body: str,
    attachments: List[Path],
) -> None:
    if not EMAIL_ENABLED:
        return

    if EMAIL_CONFIG.get("MODE") != "graph":
        raise ValueError("Only EMAIL_CONFIG['MODE'] = 'graph' is supported.")

    send_report_email_graph(
        tenant_id=EMAIL_CONFIG["GRAPH_TENANT_ID"],
        client_id=EMAIL_CONFIG["GRAPH_CLIENT_ID"],
        client_secret=EMAIL_CONFIG["GRAPH_CLIENT_SECRET"],
        sender=EMAIL_CONFIG["GRAPH_SENDER"],
        recipients=EMAIL_CONFIG["EMAIL_TO"],
        subject=subject,
        body_html=body,
        attachments=attachments,
    )


def run_scrapers(
    scrapers: List[object],
    existing_jobs: Dict[Tuple[str, str], Dict[str, str]],
    scraped_date: str,
    title_blacklist,
    classifier: JobClassifier,
) -> tuple[
    Dict[Tuple[str, str], Dict[str, str]],
    List[tuple[str, str]],
    List[Dict[str, str]],
    List[Dict[str, str]],
]:
    broken_scrapers: List[tuple[str, str]] = []
    all_new_jobs: List[Dict[str, str]] = []
    all_removed_jobs: List[Dict[str, str]] = []

    for scraper in scrapers:
        scraper_name = getattr(scraper, "source", scraper.__class__.__name__)

        try:
            scraped_jobs = scraper.scrape_jobs()
            scraped_jobs = filter_scraped_jobs(scraped_jobs, title_blacklist, classifier)
            existing_jobs, new_jobs, removed_jobs = merge_scrape_results(
                existing_jobs,
                scraped_jobs,
                scraped_date,
            )
            all_new_jobs.extend(new_jobs)
            all_removed_jobs.extend(removed_jobs)

        except Exception as exc:
            error_message = f"{type(exc).__name__}: {exc}"
            broken_scrapers.append((scraper_name, error_message))
            traceback.print_exc()

    return existing_jobs, broken_scrapers, all_new_jobs, all_removed_jobs


def main() -> None:
    scraped_date = date.today().isoformat()
    title_blacklist = load_title_blacklist(TITLE_BLACKLIST_FILE)

    classifier = JobClassifier(
        category_rules_file=CATEGORY_RULES_FILE,
        title_overrides_file=TITLE_OVERRIDES_FILE,
    )

    scrapers = [
        ArdisScraper(),
        AstridScraper(),
        BattMobilityScraper(),
        BizzTalentScraper(),
        CheopsScraper(),
        CobralScraper(),
        ColruytGroupScraper(),
        CrelanScraper(),
        DataCleanScraper(),
        DecathlonScraper(),
        District09Scraper(),
        DurabrikScraper(),
        EqualMindsScraper(),
        EvaraScraper(),
        FluxysScraper(),
        HarveyNashScraper(),
        # HaysScraper(),
        HieronymusScraper(),
        HOGentScraper(),
        HRPlanetScraper(),
        InfrabelScraper(),
        ITPlanetScraper(),
        ITProviderScraper(),
        ITSGroupScraper(),
        LogiTechnicScraper(),
        MeatAndMoreScraper(),
        OostVlaanderenScraper(),
        ProjinitScraper(),
        ResolvusScraper(),
        SignpostScraper(),
        SimacScraper(),
        SolidarisScraper(),
        SplendITScraper(),
        # SynamediaScraper(),
        TaborGroepScraper(),
        TalenciaScraper(),
        TelevicScraper(),
        UilenspelScraper(),
        UpgradeEstateScraper(),
        VDKScraper(),
        VerdonScraper(),
        VlaanderenConnectScraper(),
        VolvoCarsScraper(),
        XCareScraper(),
        XelorScraper(),
    ]

    existing_jobs = load_existing_jobs(CSV_FILE, classifier)

    if not DISABLE_BLACKLIST:
        existing_jobs = remove_blacklisted_existing_jobs(existing_jobs, title_blacklist)

    broken_scrapers: List[tuple[str, str]] = []
    new_jobs_this_run: List[Dict[str, str]] = []
    removed_jobs_this_run: List[Dict[str, str]] = []

    if RUN_SCRAPERS:
        existing_jobs, broken_scrapers, new_jobs_this_run, removed_jobs_this_run = run_scrapers(
            scrapers=scrapers,
            existing_jobs=existing_jobs,
            scraped_date=scraped_date,
            title_blacklist=title_blacklist,
            classifier=classifier,
        )

    active_jobs = get_active_jobs(existing_jobs)

    save_jobs_csv(CSV_FILE, active_jobs)
    save_jobs_excel(EXCEL_FILE, active_jobs)
    save_unrelated_jobs_txt(UNRELATED_JOBS_FILE, active_jobs)
    save_broken_scrapers_txt(BROKEN_SCRAPERS_FILE, broken_scrapers)
    save_removed_jobs_txt(REMOVED_JOBS_FILE, removed_jobs_this_run)

    category_counts = build_category_counts(active_jobs)
    unrelated_count = count_unrelated_jobs(active_jobs)

    email_body = build_email_body(
        category_counts=category_counts,
        new_jobs=new_jobs_this_run,
        removed_jobs=removed_jobs_this_run,
        unrelated_count=unrelated_count,
        broken_scrapers=broken_scrapers,
        scraped_date=scraped_date,
    )

    attachments = [
        CSV_FILE,
        EXCEL_FILE,
        UNRELATED_JOBS_FILE,
        BROKEN_SCRAPERS_FILE,
        REMOVED_JOBS_FILE,
    ]

    send_report_email(
        subject=f"Job scraper report - {scraped_date}",
        body=email_body,
        attachments=attachments,
    )


if __name__ == "__main__":
    main()